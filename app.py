from flask import Flask, render_template, request, send_file, redirect, url_for
import os
import pandas as pd
import time
from openpyxl.styles import Font

# 🔹 LOGIC IMPORTS
from logic.excel_reader import read_excel_safely
from logic.calculations import calculate_settlement
from logic.order_profit import generate_order_profit
from logic.order_mismatch import generate_order_mismatch

app = Flask(__name__)

# ================= PATHS =================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, "uploads")

# ✅ Render / server crash fix
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

print("🔥 APP FILE LOADED FROM:", __file__)

# ================= HOME =================
@app.route("/")
def home():
    return render_template("index.html")

# ================= UPLOAD =================
@app.route("/upload/<platform>", methods=["GET", "POST"])
def upload(platform):
    platform = platform.lower()
    platform_folder = os.path.join(UPLOAD_FOLDER, platform)
    os.makedirs(platform_folder, exist_ok=True)

    if request.method == "POST":
        order_file = request.files.get("order_file")
        payment_file = request.files.get("payment_file")

        if not order_file or not payment_file:
            return "❌ Order & Payment files both required"

        order_file.save(os.path.join(platform_folder, "order.xlsx"))
        payment_file.save(os.path.join(platform_folder, "payment.xlsx"))

        # ✅ upload ke baad dashboard
        return redirect(url_for("dashboard", platform=platform))

    return render_template("upload.html", platform=platform)

# ================= DASHBOARD =================
@app.route("/dashboard/<platform>")
def dashboard(platform):
    platform = platform.lower()
    payment_path = os.path.join(UPLOAD_FOLDER, platform, "payment.xlsx")

    # ❌ payment file missing
    if not os.path.exists(payment_path):
        return render_template(
            "dashboard.html",
            platform=platform,
            data=None,
            error="❌ Payment file not found. Please upload again."
        )

    try:
        payment_df = read_excel_safely(payment_path)
    except Exception as e:
        return render_template(
            "dashboard.html",
            platform=platform,
            data=None,
            error=str(e)
        )

    # ❌ empty sheet
    if payment_df is None or payment_df.empty:
        return render_template(
            "dashboard.html",
            platform=platform,
            data=None,
            error="❌ Payment file is empty or invalid."
        )

    from_date = request.args.get("from")
    to_date = request.args.get("to")

    # ✅ MULTI-PLATFORM SETTLEMENT LOGIC
    data = calculate_settlement(
        payment_df,
        platform=platform,
        from_date=from_date,
        to_date=to_date
    )

    return render_template(
        "dashboard.html",
        platform=platform,
        data=data,
        error=None,
        from_date=from_date,
        to_date=to_date
    )

# ================= DOWNLOAD ORDER PROFIT =================
@app.route("/download-order-profit/<platform>")
def download_order_profit(platform):
    platform = platform.lower()

    order_df = read_excel_safely(
        os.path.join(UPLOAD_FOLDER, platform, "order.xlsx")
    )
    payment_df = read_excel_safely(
        os.path.join(UPLOAD_FOLDER, platform, "payment.xlsx")
    )

    profit_df = generate_order_profit(order_df, payment_df)

    ts = time.strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(
        UPLOAD_FOLDER, platform, f"order_profit_{ts}.xlsx"
    )

    profit_df.to_excel(output_path, index=False)

    return send_file(
        output_path,
        as_attachment=True,
        download_name=f"{platform}_order_profit.xlsx"
    )

# ================= DOWNLOAD ORDER MISMATCH =================
@app.route("/download-order-mismatch/<platform>")
def download_order_mismatch(platform):
    platform = platform.lower()

    order_df = read_excel_safely(
        os.path.join(UPLOAD_FOLDER, platform, "order.xlsx")
    )
    payment_df = read_excel_safely(
        os.path.join(UPLOAD_FOLDER, platform, "payment.xlsx")
    )

    mismatch_df = generate_order_mismatch(order_df, payment_df)

    ts = time.strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(
        UPLOAD_FOLDER, platform, f"order_mismatch_{ts}.xlsx"
    )

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        mismatch_df.to_excel(writer, index=False, sheet_name="Mismatch")
        ws = writer.book["Mismatch"]

        red_font = Font(color="FF0000", bold=True)
        yellow_font = Font(color="FFC000", bold=True)

        status_col = mismatch_df.columns.get_loc("Status") + 1

        for row in range(2, ws.max_row + 1):
            status = ws.cell(row=row, column=status_col).value

            if status == "UNPAID":
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).font = red_font

            elif status == "PAID (ORDER NOT FOUND)":
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).font = yellow_font

    return send_file(
        output_path,
        as_attachment=True,
        download_name=f"{platform}_order_mismatch.xlsx"
    )

# ================= RUN =================
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 10000))
    app.run(host="0.0.0.0", port=port)
